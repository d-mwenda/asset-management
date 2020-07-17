from django.shortcuts import render, render_to_response, HttpResponseRedirect
from django.views.generic import ListView, FormView, CreateView, UpdateView, DetailView
from django.db import transaction
from django.urls import reverse_lazy
from django.http import HttpResponse, Http404
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin

# third party libraries imports
from rest_framework import generics
from rest_framework.viewsets import ModelViewSet
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# asset-management imports
from .models import Assets, AssetUsersDetails, AssetTypes, AssetMakes, AssetModels, Vendors, AssetOwners, Offices,\
    AssetIssuanceRegister, NonHumanUsers
from .forms import AddAssetForm, AddAssetTypeForm, AddAssetModelForm, AddAssetMakeForm, AddVendorForm,\
    AddAssetOwnerForm, AddOfficeLocationForm, AssetIssuanceForm, SearchForm, AssetsExcelExportForm, \
    AssetWithdrawalForm
from .serializers import AssetTypesSerializer, AssetMakesSerializer, AssetModelsSerializer, VendorsSerializer, \
    OfficesSerializer, AssetOwnersSerializer, AssetUsersDetailsSerializer, NonHumanUsersSerializer, AssetsSerializer, \
    AssetIssuanceRegisterSerializer
from toolbox.dbqueries import get_all_items, check_if_queryset

# TODO add support for exporting to xlsx. Can be a module
# TODO add support for batch uploading from xlsx and/or csv


class AssetTypesViewSet(ModelViewSet):
    """
    This class offers an API endpoint for operations on all Asset Types in the database.
    """
    queryset = AssetTypes.objects.all()
    serializer_class = AssetTypesSerializer


class AssetMakesViewSet(ModelViewSet):
    """
    This class offers an API endpoint for operations on all Asset Makes in the database.
    """
    queryset = AssetMakes.objects.all()
    serializer_class = AssetMakesSerializer


class AssetModelsViewSet(ModelViewSet):
    """
    This class offers an API endpoint for operations on all Asset Models in the database.
    """
    queryset = AssetModels.objects.all()
    serializer_class = AssetModelsSerializer


class VendorsViewSet(ModelViewSet):
    """
    This class offers an API endpoint for operations on all Asset Vendors in the database.
    """
    queryset = Vendors.objects.all()
    serializer_class = VendorsSerializer


class OfficesViewSet(ModelViewSet):
    """
    This class offers an API endpoint for operations on all Asset Models in the database.
    """
    queryset = Offices.objects.all()
    serializer_class = OfficesSerializer


class AssetOwnersViewSet(ModelViewSet):
    """
    This class offers an API endpoint for operations on all Asset Models in the database.
    """
    queryset = AssetOwners.objects.all()
    serializer_class = AssetOwnersSerializer


class AssetUsersDetailsViewSet(ModelViewSet):
    """
    This class offers an API endpoint for operations on all Asset Models in the database.
    """
    queryset = AssetUsersDetails.objects.all()
    serializer_class = AssetUsersDetailsSerializer


class NonHumanUsersViewSet(ModelViewSet):
    """
    This class offers an API endpoint for operations on all Asset Models in the database.
    """
    queryset = NonHumanUsers.objects.all()
    serializer_class = NonHumanUsersSerializer


class AssetsViewSet(ModelViewSet):
    """
    This class offers an API endpoint for operations on all Assets in the database.
    """
    queryset = Assets.objects.all()
    serializer_class = AssetsSerializer


class AssetIssuanceRegisterViewSet(ModelViewSet):
    """
    This class offers an API endpoint for operations on all Asset Models in the database.
    """
    queryset = AssetIssuanceRegister.objects.all()
    serializer_class = AssetIssuanceRegisterSerializer


class ListAssets(LoginRequiredMixin, ListView):
    # This class lists all the assets in the register, regardless of their status
    # TODO probably in future this needs to be limited to only active / undisposed assets
    # TODO users assigned should appear in the list
    template_name = 'assets_register/view_assets.html'
    login_url = ''
    redirect_field_name = ''
    context_object_name = 'asset_register'
    model = Assets
    form = AssetsExcelExportForm()

    def export_xlsx(self, relevant_fields, queryset, *args, **kwargs):
        """
        export all assets to xlsx
        Todo: consider using this method as an util for all excel exports
        """
        asset_list = Assets.objects.all().values(*relevant_fields)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="asset register.xlsx"'
        asset_register = Workbook()
        worksheet = asset_register.active
        worksheet.title = "All Assets"
        row_num = 1
        for col_num, column_title in enumerate(relevant_fields, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
        for asset in asset_list:
            row_num += 1
            row = [
                asset[attribute] for attribute in asset
            ]

            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        asset_register.save(response)
        return response

    def get_context_data(self, *, object_list=None, **kwargs):
        ctx = super(ListAssets, self).get_context_data()
        ctx['view_name'] = 'View Assets'
        ctx['form'] = self.form
        return ctx

    def form_valid(self,request, form):
        relevant_fields = []
        clean_form = form.cleaned_data
        for field in clean_form:
            if clean_form[field]:
                relevant_fields.append(field)
        return self.export_xlsx(relevant_fields)

    def form_invalid(self, form, request):
        # include error message
        return self.render_to_response(self.get_context_data(form))

    def post(self, request, *args, **kwargs):
        form = AssetsExcelExportForm(request.POST)
        if form.is_valid():
            return self.form_valid(request, form)
        else:
            return self.form_invalid(form)


class AssetStore(LoginRequiredMixin, ListView):
    template_name = "assets_register/view_assets.html"
    model = Assets
    context_object_name = "asset_register"
    view_name = "View Assets in the Store"

    def get_context_data(self, *, object_list=None, **kwargs):
        ctx = super(AssetStore, self).get_context_data()
        ctx['view_name'] = self.view_name
        return ctx

    def get_queryset(self):
        queryset = super(AssetStore, self).get_queryset()
        return queryset.filter(status_of_usage='uis')


class UserAssignedAssets(LoginRequiredMixin, ListView):
    template_name = "assets_register/user_assigned_assets.html"
    model = AssetIssuanceRegister
    pk_url_kwarg = "user_id"
    context_object_name = "asset_register"
    view_name = "User Assigned Assets"

    def get_queryset(self):
        """
        From the queryset, filter the assets assigned to the user in question
        TODO: consider ordering by returned so that the returned ones appear at the bottom
        """
        queryset = super(UserAssignedAssets, self).get_queryset()
        return queryset.filter(user=self.kwargs['user_id'])

    def get_user(self):
        return AssetUsersDetails.objects.get(id=self.kwargs['user_id'])

    def get_context_data(self, *, object_list=None, **kwargs):
        ctx = super(UserAssignedAssets, self).get_context_data()
        ctx['view_name'] = self.view_name
        ctx['user'] = self.get_user()
        return ctx


class AssetProfile(LoginRequiredMixin, DetailView):
    template_name = "assets_register/asset_profile.html"
    model = Assets
    context_object_name = "asset"
    pk_url_kwarg = "asset_id"
    view_name = 'Asset Profile'

    def get_usage_queryset(self):
        return AssetIssuanceRegister.objects.filter(asset=self.kwargs['asset_id'])

    def get_context_data(self, **kwargs):
        ctx = super(AssetProfile, self).get_context_data(**kwargs)
        ctx['usage_history'] = self.get_usage_queryset()
        ctx['view_name'] = self.view_name
        return ctx


class SearchAsset(LoginRequiredMixin, FormView):
    # This class provides both simple and advanced search in the asset database
    # TODO: consider using an already existing strong search engine from pip
    template_name = 'assets_register/search_asset.html'
    context_data = dict()
    context_data['view_name'] = 'Search Assets'

    def get(self, request, *args, **kwargs):
        context_data = self.context_data
        search_form = SearchForm(request.GET)
        if search_form.is_valid():
            query = search_form.cleaned_data['query']
            result = Assets.objects.filter(serial_number=query)
            context_data['result'] = result
        context_data['search_form'] = search_form
        return render(request, self.template_name, context_data)

    def post(self, request, *args, **kwargs):
        pass


class AddAsset(LoginRequiredMixin, FormView):
    template_name = 'assets_register/add_asset_metadata.html'
    context_data = dict()
    context_data['view_name'] = 'Add Asset'

    def get(self, request, *args, **kwargs):
        context_data = self.context_data
        context_data['form'] = AddAssetForm()
        context_data['cancel_url'] = reverse_lazy('u_list_assets')
        return render(request, self.template_name, context_data)

    def post(self, request, *args, **kwargs):
        context_data = self.context_data
        add_asset_form = AddAssetForm(request.POST)
        if add_asset_form.is_valid():
            add_asset_form.save()
            context_data['message'] = 'Asset saved successfully.'
        else:
            context_data['error_message'] = 'Form is not validly completed'
            context_data['form'] = add_asset_form
        return render(request, self.template_name, context_data)


class IssueAsset(LoginRequiredMixin, CreateView):
    # This class handles the functionality of issuing out assets
    # Todo: send an email to the user after the asset issued to them
    template_name = 'assets_register/issue_asset.html'
    form_class = AssetIssuanceForm
    model = AssetIssuanceRegister
    context_object_name = 'asset'
    pk_url_kwarg = 'asset'
    success_url = reverse_lazy('u_list_assets')
    cancel_url = reverse_lazy('u_list_assets')
    already_in_use = None

    def get_queryset(self):
        #  This is the asset to be issued out
        return Assets.objects.filter(id=self.kwargs['asset'])

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        asset_issuance_form = AssetIssuanceForm(request.POST)
        if asset_issuance_form.is_valid():
            return self.form_valid(asset_issuance_form)
        else:
            return self.form_invalid(asset_issuance_form)

    def form_valid(self, asset_issuance_form):
        candidate_asset = asset_issuance_form.cleaned_data['asset']
        asset = Assets.objects.get(id=candidate_asset.id)
        if self.check_asset_is_free(candidate_asset):
            with transaction.atomic():
                asset.status_of_usage = "iss"
                asset.save()
                asset_issuance_form.save()
            return HttpResponseRedirect(self.get_success_url())
        else:
            return self.render_to_response(
                {"message": "already issued out"}
            )

    def check_asset_is_free(self, candidate_asset):
        """
        This method checks that the asset is not issued to avoid duplicate issuance.
        If it's issued, the current user is returned
        """
        asset = Assets.objects.get(id=candidate_asset.id)
        if asset.status_of_usage == "iss":
            return False
        else:
            return True

    def get_asset_user(self, candidate_asset):
        """
        This method retrieves the user assigned a certain asset
        """
        return AssetIssuanceRegister.objects.filter(asset=candidate_asset.id).filter(returned=False)

    def form_invalid(self, form):
       return self.render_to_response(self.get_context_data(form=form, success='form is invalid')) # debug

    def get_context_data(self, **kwargs):
        context = super(IssueAsset, self).get_context_data(**kwargs)
        context['view_name'] = 'Issue Out Asset'
        context['cancel_url'] = self.cancel_url
        context['asset'] = self.get_queryset()
        return context

    def get_initial(self):
        asset = self.get_object(self.queryset)
        return {
            'asset': asset,
        }


class AssetWithdrawal(LoginRequiredMixin, UpdateView):
    template_name = 'assets_register/withdraw_asset.html'
    model = AssetIssuanceRegister
    form_class = AssetWithdrawalForm
    context_object_name = 'asset'
    pk_url_kwarg = 'asset'
    success_url = reverse_lazy('u_list_assets')
    cancel_url = reverse_lazy('u_list_assets')

    def get_object(self, queryset=None):
        """
        get the asset issuance record to be withdrawn from use.
        """
        if queryset is None:
            queryset = self.get_queryset()
            queryset = queryset.filter(asset=self.kwargs['asset']).filter(returned=False)

            try:
                # Get the single item from the filtered queryset
                obj = queryset.get()
            except queryset.model.DoesNotExist:
                raise Http404(("No %(verbose_name)s found matching the query") %
                              {'verbose_name': queryset.model._meta.verbose_name})
            return obj

    def get_context_data(self, **kwargs):
        context = super(AssetWithdrawal, self).get_context_data(**kwargs)
        context['view_name'] = 'Withdraw Asset'
        context['cancel_url'] = self.cancel_url
        context['issuance_record'] = self.get_object()
        return context

    def form_valid(self, form):
        with transaction.atomic():
            asset = Assets.objects.get(id=form.cleaned_data['asset'].id)
            asset.status_of_usage = 'uis'
            asset.save()
            issuance_record = self.model.objects.filter(asset=asset).filter(returned=False)
            issuance_record = issuance_record.get()
            issuance_record.date_returned = form.cleaned_data['date_returned']
            issuance_record.comment_on_return = form.cleaned_data['comment_on_return']
            issuance_record.returned = True
            issuance_record.save()
        return HttpResponseRedirect(self.get_success_url())


class AssetsUsers(LoginRequiredMixin, ListView):
    template_name = 'assets_register/asset_users.html'
    context_data = dict()
    context_data['view_name'] = 'Assets Users'

    # def get(self, request, *args, **kwargs):
    #     context_data = self.context_data
    #     # asset_users = get_all_items(AssetUsers) todo debug this
    #     if check_if_queryset(asset_users):
    #         context_data['asset_users'] = asset_users
    #     else:
    #         context_data['no_asset_users'] = asset_users
    #     return render(request, self.template_name, context_data)

    def post(self, request, *args, **kwargs):
        pass


class AssetsAdministration(LoginRequiredMixin, ListView):
    # Todo: depreciate this class for redundancy after confirming it's no longer in use
    template_name = 'assets_register/assets_administration.html'
    context_data = dict()
    context_data['view_name'] = 'Assets Administration'

    def get(self, request, *args, **kwargs):
        context_data = self.context_data
        asset_types = get_all_items(AssetTypes)
        asset_makes = get_all_items(AssetMakes)
        asset_models = get_all_items(AssetModels)
        asset_vendors = get_all_items(Vendors)
        asset_owners = get_all_items(AssetOwners)
        offices = get_all_items(Offices)
        if check_if_queryset(asset_types):
            context_data['asset_types'] = asset_types
        else:
            context_data['no_asset_types'] = asset_types
        if check_if_queryset(asset_makes):
            context_data['asset_makes'] = asset_makes
        else:
            context_data['no_asset_makes'] = asset_makes
        if check_if_queryset(asset_models):
            context_data['asset_models'] = asset_models
        else:
            context_data['no_asset_models'] = asset_models
        if check_if_queryset(asset_vendors):
            context_data['asset_vendors'] = asset_vendors
        else:
            context_data['no_asset_vendors'] = asset_vendors
        if check_if_queryset(asset_owners):
            context_data['asset_owners'] = asset_owners
        else:
            context_data['no_asset_owners'] = asset_owners
        if check_if_queryset(offices):
            context_data['offices'] = offices
        else:
            context_data['no_offices'] = offices
        return render(request, self.template_name, context_data)

    def post(self, request, *args, **kwargs):
        pass


# The views below perform the functionality of adding asset metadata


class AddAssetType(LoginRequiredMixin, FormView):
    template_name = 'assets_register/add_asset_metadata.html'
    context_data = dict()
    context_data['view_name'] = 'Add Asset Type'

    def get(self, request, *args, **kwargs):
        context_data = self.context_data
        context_data['form'] = AddAssetTypeForm()
        context_data['cancel_url'] = reverse_lazy('u_view_asset_types')
        return render(request, self.template_name, context_data)

    def post(self, request, *args, **kwargs):
        context_data = self.context_data
        add_asset_type_form = AddAssetTypeForm(request.POST)
        if add_asset_type_form.is_valid():
            add_asset_type_form.save()
            context_data['message'] = 'Asset type saved successfully.'
        else:
            context_data['error_message'] = 'Form is not validly completed'
        return render(request, self.template_name, context_data)


class AddAssetModel(LoginRequiredMixin, FormView):
    template_name = 'assets_register/add_asset_metadata.html'
    context_data = dict()
    context_data['view_name'] = 'Add Asset Model'

    def get(self, request, *args, **kwargs):
        context_data = self.context_data
        context_data['form'] = AddAssetModelForm()
        context_data['cancel_url'] = reverse_lazy('u_view_asset_models')
        return render(request, self.template_name, context_data)

    def post(self, request, *args, **kwargs):
        context_data = self.context_data
        add_asset_model_form = AddAssetModelForm(request.POST)
        if add_asset_model_form.is_valid():
            add_asset_model_form.save()
            context_data['message'] = 'Asset model saved successfully.'
        else:
            context_data['error_message'] = 'Form is not validly completed'
        return render(request, self.template_name, context_data)


class AddAssetMake(LoginRequiredMixin, FormView):
    template_name = 'assets_register/add_asset_metadata.html'
    context_data = dict()
    context_data['view_name'] = 'Add Asset Make'

    def get(self, request, *args, **kwargs):
        context_data = self.context_data
        context_data['form'] = AddAssetMakeForm()
        context_data['cancel_url'] = reverse_lazy('u_view_asset_makes')
        return render(request, self.template_name, context_data)

    def post(self, request, *args, **kwargs):
        context_data = self.context_data
        add_asset_make_form = AddAssetMakeForm(request.POST)
        if add_asset_make_form.is_valid():
            add_asset_make_form.save()
            context_data['message'] = 'Asset make saved successfully.'
        else:
            context_data['error_message'] = 'Form is not validly completed'
        return render(request, self.template_name, context_data)


class AddAssetVendor(LoginRequiredMixin, FormView):
    template_name = 'assets_register/add_asset_metadata.html'
    context_data = dict()
    context_data['view_name'] = 'Add Vendor'

    def get(self, request, *args, **kwargs):
        context_data = self.context_data
        context_data['form'] = AddVendorForm()
        context_data['cancel_url'] = reverse_lazy('u_view_asset_vendors')
        return render(request, self.template_name, context_data)

    def post(self, request, *args, **kwargs):
        context_data = self.context_data
        add_vendor_form = AddVendorForm(request.POST)
        if add_vendor_form.is_valid():
            add_vendor_form.save()
            context_data['message'] = 'Vendor saved successfully.'
        else:
            context_data['error_message'] = 'Form is not validly completed'
        return render(request, self.template_name, context_data)


class AddAssetOwner(LoginRequiredMixin, FormView):
    template_name = 'assets_register/add_asset_metadata.html'
    context_data = dict()
    context_data['view_name'] = 'Add Asset Owner'

    def get(self, request, *args, **kwargs):
        context_data = self.context_data
        context_data['form'] = AddAssetOwnerForm
        context_data['cancel_url'] = reverse_lazy('u_view_asset_owners')
        return render(request, self.template_name, context_data)

    def post(self, request, *args, **kwargs):
        context_data = self.context_data
        add_asset_owner_form = AddAssetOwnerForm(request.POST)
        if add_asset_owner_form.is_valid():
            add_asset_owner_form.save()
            context_data['message'] = 'Asset owner saved successfully.'
        else:
            context_data['error_message'] = 'Form is not validly completed'
        return render(request, self.template_name, context_data)


class AddOfficeLocation(LoginRequiredMixin, FormView):
    template_name = 'assets_register/add_asset_metadata.html'
    context_data = dict()
    context_data['view_name'] = 'Add Office Location'

    def get(self, request, *args, **kwargs):
        context_data = self.context_data
        context_data['form'] = AddOfficeLocationForm
        context_data['cancel_url'] = reverse_lazy('u_view_offices')
        return render(request, self.template_name, context_data)

    def post(self, request, *args, **kwargs):
        context_data = self.context_data
        add_office_location_form = AddOfficeLocationForm(request.POST)
        if add_office_location_form.is_valid():
            add_office_location_form.save()
            context_data['message'] = 'Office location saved successfully.'
        else:
            context_data['error_message'] = 'Form is not validly completed'
        return render(request, self.template_name, context_data)


class AddAssetUser(LoginRequiredMixin, FormView):
    # This class is used to add people to whom assets will be issued.
    # Todo consider managing users using a standalone app that has even the ability of integration with identity systems
    template_name = 'assets_register/add_asset_metadata.html'
    context_data = dict()
    context_data['view_name'] = 'Add Asset User'

    def get(self, request, *args, **kwargs):
        context_data = self.context_data
        # Todo: the below url is just a placeholder to prevent errors. create a list user view
        context_data['cancel_url'] = reverse_lazy('u_list_assets')
        # context_data['form'] = AddAssetUserForm
        return render(request, self.template_name, context_data)

    # def post(self, request, *args, **kwargs):
    #     context_data = self.context_data
    #     # add_asset_user_form = AddAssetUserForm(request.POST)
    #     # if add_asset_user_form.is_valid():
    #     #     add_asset_user_form.save()
    #     #     context_data['message'] = 'Asset user saved successfully.'
    #     else:
    #         context_data['error_message'] = 'Form is not validly completed'
    #     return render(request, self.template_name, context_data)


# The views below handle the listing of various asset management parameters


class ViewAssetTypes(LoginRequiredMixin, ListView):
    # This class lists all the available asset types in the database
    template_name = 'assets_register/view_asset_types.html'
    context_object_name = 'asset_types'
    model = AssetTypes

    def get_context_data(self, object_list=None, *args, **kwargs):
        ctx = super(ViewAssetTypes, self).get_context_data()
        ctx['view_name'] = 'View Asset Types'
        return ctx


class ViewAssetMakes(LoginRequiredMixin, ListView):
    # This class lists all the available asset makes in the database
    template_name = 'assets_register/view_asset_makes.html'
    context_object_name = 'asset_makes'
    model = AssetMakes

    def get_context_data(self, object_list=None, *args, **kwargs):
        ctx = super(ViewAssetMakes, self).get_context_data()
        ctx['view_name'] = 'View Asset Makes'
        return ctx


class ViewAssetModels(LoginRequiredMixin, ListView):
    template_name = 'assets_register/view_asset_models.html'
    context_object_name = 'asset_models'
    model = AssetModels

    def get_context_data(self, object_list=None, *args, **kwargs):
        ctx = super(ViewAssetModels, self).get_context_data()
        ctx['view_name'] = 'View Asset Models'
        return ctx


class ViewAssetOwners(LoginRequiredMixin, ListView):
    template_name = 'assets_register/view_asset_owners.html'
    context_object_name = 'asset_owners'
    model = AssetOwners

    def get_context_data(self, object_list=None, *args, **kwargs):
        ctx = super(ViewAssetOwners, self).get_context_data()
        ctx['view_name'] = 'View Asset Owners'
        return ctx


class ViewVendors(LoginRequiredMixin, ListView):
    template_name = 'assets_register/view_vendors.html'
    context_object_name = 'asset_vendors'
    model = Vendors

    def get_context_data(self, object_list=None, *args, **kwargs):
        ctx = super(ViewVendors, self).get_context_data()
        ctx['view_name'] = 'View Asset Vendors'
        return ctx


class ViewOfficeLocations(LoginRequiredMixin, ListView):
    template_name = 'assets_register/view_office_locations.html'
    context_object_name = 'offices'
    model = Offices

    def get_context_data(self, object_list=None, *args, **kwargs):
        ctx = super(ViewOfficeLocations, self).get_context_data()
        ctx['view_name'] = 'View Office Locations'
        return ctx


# The views below handle the update functionality of the various asset management parameters

class UpdateAssetType(LoginRequiredMixin, UpdateView):
    # This class is used to edit the asset types
    template_name = 'assets_register/add_asset_metadata.html'
    context_object_name = 'asset_type'
    model = AssetTypes
    form_class = AddAssetTypeForm
    pk_url_kwarg = 'asset_type'
    cancel_url = reverse_lazy('u_view_asset_types')
    success_url = reverse_lazy('u_view_asset_types')

    def get_context_data(self, object_list=None, *args, **kwargs):
        ctx = super(UpdateAssetType, self).get_context_data()
        ctx['view_name'] = 'Edit Asset Type'
        ctx['cancel_url'] = self.cancel_url
        ctx['warning'] = 'Take care! Modifying an asset type alters the asset type of existing assets with this type!!!'
        return ctx


class UpdateAssetMake(LoginRequiredMixin, UpdateView):
    # This class is used to edit the asset makes
    template_name = 'assets_register/add_asset_metadata.html'
    context_object_name = 'asset_make'
    model = AssetMakes
    form_class = AddAssetMakeForm
    pk_url_kwarg = 'asset_make'
    cancel_url = reverse_lazy('u_view_asset_makes')
    success_url = reverse_lazy('u_view_asset_types')

    def get_context_data(self, object_list=None, *args, **kwargs):
        ctx = super(UpdateAssetMake, self).get_context_data()
        ctx['view_name'] = 'Edit Asset Make'
        ctx['cancel_url'] = self.cancel_url
        ctx['warning'] = 'Take care! Modifying an asset make alters the asset make of existing assets with this make!!!'
        return ctx


class UpdateAssetModel(LoginRequiredMixin, UpdateView):
    # This class is used to edit the asset models
    template_name = 'assets_register/add_asset_metadata.html'
    context_object_name = 'asset_model'
    model = AssetModels
    form_class = AddAssetModelForm
    pk_url_kwarg = 'asset_model'
    cancel_url = reverse_lazy('u_view_asset_models')
    success_url = reverse_lazy('u_view_asset_models')

    def get_context_data(self, object_list=None, *args, **kwargs):
        ctx = super(UpdateAssetModel, self).get_context_data()
        ctx['view_name'] = 'Edit Asset Model'
        ctx['cancel_url'] = self.cancel_url
        ctx['warning'] = 'Take care! Modifying an asset model alters the asset model parameters of existing assets' \
                         ' with this model!!!'
        return ctx


class UpdateAssetOwner(LoginRequiredMixin, UpdateView):
    # This class is used to edit the asset owners
    template_name = 'assets_register/add_asset_metadata.html'
    context_object_name = 'asset_owner'
    model = AssetOwners
    form_class = AddAssetOwnerForm
    pk_url_kwarg = 'asset_owner'
    cancel_url = reverse_lazy('u_view_asset_owners')
    success_url = reverse_lazy('u_view_asset_owners')

    def get_context_data(self, object_list=None, *args, **kwargs):
        ctx = super(UpdateAssetOwner, self).get_context_data()
        ctx['view_name'] = 'Edit Asset Owner'
        ctx['cancel_url'] = self.cancel_url
        ctx['warning'] = 'Take care! Modifying an asset owner alters the asset owner details of existing assets' \
                         ' with this owner!!!'
        return ctx


class UpdateAssetVendor(LoginRequiredMixin, UpdateView):
    # This class is used to edit the asset models
    template_name = 'assets_register/add_asset_metadata.html'
    context_object_name = 'asset_vendor'
    model = Vendors
    form_class = AddVendorForm
    pk_url_kwarg = 'asset_vendor'
    cancel_url = reverse_lazy('u_view_asset_vendors')
    success_url = reverse_lazy('u_view_asset_vendors')

    def get_context_data(self, object_list=None, *args, **kwargs):
        ctx = super(UpdateAssetVendor, self).get_context_data()
        ctx['view_name'] = 'Edit Vendor'
        ctx['cancel_url'] = self.cancel_url
        ctx['warning'] = 'Take care! Modifying a vendor alters the vendor details of existing assets with this vendor!!!'
        return ctx


class UpdateOfficeLocation(LoginRequiredMixin, UpdateView):
    # This class is used to edit the asset models
    template_name = 'assets_register/add_asset_metadata.html'
    context_object_name = 'asset_location'
    model = Offices
    form_class = AddOfficeLocationForm
    pk_url_kwarg = 'office'
    cancel_url = reverse_lazy('u_view_offices')
    success_url = reverse_lazy('u_view_offices')

    def get_context_data(self, object_list=None, *args, **kwargs):
        ctx = super(UpdateOfficeLocation, self).get_context_data()
        ctx['view_name'] = 'Edit Office Location'
        ctx['cancel_url'] = self.cancel_url
        ctx['warning'] = 'Take care! Modifying an office location alters the office location of existing assets ' \
                         'with this office location!!!'
        return ctx


# Maintain the below function for backward compatibility until you're sure its usage on this view has been depreciated
def get_asset(asset):
    try:
        return Assets.objects.get(asset_number=asset)
    except (Assets.DoesNotExist, Assets.MultipleObjectsReturned):
        return "Error while trying to retrieve the asset"
